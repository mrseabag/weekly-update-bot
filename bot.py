#!/usr/bin/env python3
"""
Weekly Team Update Bot
- Sends weekly questions to registered team members every Monday
- Collects text answers + file attachments
- Stores everything in an Excel spreadsheet
- Emails the manager a report every Friday
"""

import json
import logging
import os
import smtplib
import socket
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove, Update
from telegram.ext import (
    Application,
    CommandHandler,
    ConversationHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

load_dotenv()

# Railway's containers have no IPv6 egress route, but smtp.gmail.com resolves to
# both A and AAAA records — Python tries AAAA first and fails with
# "[Errno 101] Network is unreachable". Force IPv4-only DNS resolution so SMTP
# connects over IPv4 instead.
_orig_getaddrinfo = socket.getaddrinfo


def _ipv4_only_getaddrinfo(host, port, family=0, type=0, proto=0, flags=0):
    return _orig_getaddrinfo(host, port, socket.AF_INET, type, proto, flags)


socket.getaddrinfo = _ipv4_only_getaddrinfo

# ════════════════════════════════════════════════════════════════════════════
#  CONFIGURATION  —  fill these in before running
# ════════════════════════════════════════════════════════════════════════════

BOT_TOKEN = os.getenv("BOT_TOKEN")

MANAGER_EMAIL = os.getenv("MANAGER_EMAIL")     # report goes here
SENDER_EMAIL = os.getenv("SENDER_EMAIL")        # Gmail sending it
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD")

# Weekly prompt schedule (24h clock, UTC) — Friday 11:00 Ulaanbaatar time (UTC+8) = Friday 03:00 UTC
PROMPT_DAY = 4    # 0=Monday … 6=Sunday
PROMPT_HOUR = 3
PROMPT_MINUTE = 0

# Weekly report email schedule — Friday 19:00 Ulaanbaatar time (UTC+8) = Friday 11:00 UTC
REPORT_DAY = 4    # Friday
REPORT_HOUR = 11
REPORT_MINUTE = 0

# ════════════════════════════════════════════════════════════════════════════

DATA_DIR = Path("data")
EXCEL_PATH = DATA_DIR / "weekly_updates.xlsx"
USERS_PATH = DATA_DIR / "users.json"
FILES_DIR = DATA_DIR / "submitted_files"

DATA_DIR.mkdir(parents=True, exist_ok=True)
FILES_DIR.mkdir(parents=True, exist_ok=True)

# ConversationHandler states
COMPLETED, WORKING_ON, AWAITING_FILES = range(3)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)s  %(message)s",
)
logger = logging.getLogger(__name__)


# ── User storage ─────────────────────────────────────────────────────────────

def load_users() -> dict:
    if USERS_PATH.exists():
        return json.loads(USERS_PATH.read_text())
    return {}


def save_users(users: dict) -> None:
    USERS_PATH.write_text(json.dumps(users, indent=2))


def register_user(user_id: int, name: str, username: str) -> None:
    users = load_users()
    users[str(user_id)] = {"name": name, "username": username or name}
    save_users(users)


# ── Excel storage ─────────────────────────────────────────────────────────────

HEADERS = ["Долоо хоног", "Илгээсэн огноо", "Нэр", "Хэрэглэгчийн нэр",
           "Юу хийж дуусгасан бэ?", "Одоо юу дээр ажиллаж байна вэ?", "Илгээсэн файлууд"]

def _ensure_excel() -> openpyxl.Workbook:
    if EXCEL_PATH.exists():
        return openpyxl.load_workbook(EXCEL_PATH)

    wb = Workbook()
    ws = wb.active
    ws.title = "Updates"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    col_widths = [12, 18, 20, 20, 45, 45, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    wb.save(EXCEL_PATH)
    return wb


def save_update_to_excel(user_data: dict) -> None:
    wb = _ensure_excel()
    ws = wb["Updates"]

    now = datetime.now()
    week_label = now.strftime("W%V %Y")    # e.g. W28 2026

    files_str = ", ".join(user_data.get("files", [])) or "Байхгүй"

    row = [
        week_label,
        now.strftime("%Y-%m-%d %H:%M"),
        user_data.get("name", ""),
        user_data.get("username", ""),
        user_data.get("completed", ""),
        user_data.get("working_on", ""),
        files_str,
    ]

    ws.append(row)

    # Zebra-stripe every other data row for readability
    last_row = ws.max_row
    if last_row % 2 == 0:
        fill = PatternFill("solid", fgColor="D9E1F2")
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=last_row, column=col).fill = fill

    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=last_row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(EXCEL_PATH)
    logger.info("Saved update for %s to Excel", user_data.get("name"))


# ── Email ─────────────────────────────────────────────────────────────────────

EMAIL_ERROR_MESSAGES = {
    "not_configured": "MANAGER_EMAIL, SENDER_EMAIL, SENDER_PASSWORD тохиргоогоо шалгана уу.",
    "no_data": "Тайлангийн өгөгдөл алга (weekly_updates.xlsx үүсээгүй байна — Railway дээр Volume холбогдсон эсэхийг шалгана уу).",
    "smtp_error": "Имэйл серверт холбогдоход алдаа гарлаа (Gmail нэвтрэлт эсвэл сүлжээний асуудал). Дэлгэрэнгүйг лог-оос шалгана уу.",
}


def send_weekly_email() -> tuple[bool, str]:
    if not all([MANAGER_EMAIL, SENDER_EMAIL, SENDER_PASSWORD]):
        logger.warning("Email not configured — skipping report")
        return False, "not_configured"

    if not EXCEL_PATH.exists():
        logger.warning("No Excel file yet — skipping report")
        return False, "no_data"

    week_label = datetime.now().strftime("W%V %Y")

    msg = MIMEMultipart()
    msg["From"] = SENDER_EMAIL
    msg["To"] = MANAGER_EMAIL
    msg["Subject"] = f"Weekly Team Update Report — {week_label}"

    body = (
        f"Hi,\n\nPlease find attached the weekly team update report for {week_label}.\n\n"
        "The spreadsheet contains each team member's completed work, "
        "current tasks, and any files they submitted.\n\n"
        "— Weekly Update Bot"
    )
    msg.attach(MIMEText(body, "plain"))

    with open(EXCEL_PATH, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f'attachment; filename="weekly_updates_{week_label}.xlsx"',
    )
    msg.attach(part)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(SENDER_EMAIL, SENDER_PASSWORD)
            server.sendmail(SENDER_EMAIL, MANAGER_EMAIL, msg.as_string())
        logger.info("Weekly report emailed to %s", MANAGER_EMAIL)
        return True, ""
    except Exception as e:
        logger.error("Failed to send email: %s", e)
        return False, "smtp_error"


# ── Scheduled jobs ────────────────────────────────────────────────────────────

async def job_send_weekly_prompt(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Fires every Monday — sends update request to all registered users."""
    users = load_users()
    if not users:
        logger.info("No registered users yet")
        return

    for uid in users:
        try:
            await context.bot.send_message(
                chat_id=int(uid),
                text=(
                    "👋 *Долоо хоногийн тайлан илгээх цаг боллоо!*\n\n"
                    "Энэ долоо хоногийн тайланг илгээнэ үү. "
                    "/update командыг ашиглан эхлүүлнэ үү.\n\n"
                    "/update гэж бичнэ үү."
                ),
                parse_mode="Markdown",
            )
        except Exception as e:
            logger.warning("Could not message user %s: %s", uid, e)


async def job_send_weekly_report(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Fires every Friday — emails Excel report to manager."""
    success, reason = send_weekly_email()
    # Notify admin users if configured
    users = load_users()
    for uid, info in users.items():
        if info.get("is_admin"):
            status = (
                "✅ Долоо хоногийн тайлан менежерт имэйлээр илгээгдлээ."
                if success
                else f"⚠️ Имэйл илгээхэд алдаа гарлаа — {EMAIL_ERROR_MESSAGES.get(reason, 'тодорхойгүй алдаа.')}"
            )
            try:
                await context.bot.send_message(chat_id=int(uid), text=status)
            except Exception:
                pass


# ── Conversation: /update ─────────────────────────────────────────────────────

async def cmd_update(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    register_user(user.id, user.full_name, user.username)
    context.user_data.clear()
    context.user_data["name"] = user.full_name
    context.user_data["username"] = user.username or user.full_name
    context.user_data["files"] = []

    await update.message.reply_text(
        "📋 *Долоо хоногийн тайлан — 1-р асуулт (3-аас)*\n\n"
        "Энэ долоо хоногт та юу *хийж дуусгасан* бэ?\n\n"
        "_Хариултаа доор бичнэ үү._",
        parse_mode="Markdown",
    )
    return COMPLETED


async def receive_completed(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["completed"] = update.message.text
    await update.message.reply_text(
        "✅ Ойлголоо!\n\n"
        "📋 *2-р асуулт (3-аас)*\n\n"
        "Одоо та юу дээр *ажиллаж байна* вэ?\n\n"
        "_Хариултаа доор бичнэ үү._",
        parse_mode="Markdown",
    )
    return WORKING_ON


async def receive_working_on(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data["working_on"] = update.message.text

    keyboard = [["Дууслаа — файл байхгүй"]]
    await update.message.reply_text(
        "✅ Ойлголоо!\n\n"
        "📋 *3-р асуулт (3-аас)*\n\n"
        "Илгээх *файл байна уу*? "
        "(баримт бичиг, тайлан, хүснэгт, зураг гэх мэт)\n\n"
        "Нэг нэгээр нь илгээнэ үү. Дуусвал доорх товчийг дарна уу "
        "эсвэл *дууслаа* гэж бичнэ үү.",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True),
    )
    return AWAITING_FILES


async def receive_file(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Handle a file attachment during the update conversation."""
    msg = update.message
    week_label = datetime.now().strftime("W%V_%Y")
    user_slug = context.user_data.get("username", "user").replace(" ", "_")
    save_subdir = FILES_DIR / week_label / user_slug
    save_subdir.mkdir(parents=True, exist_ok=True)

    # Determine the Telegram file object
    tg_file = None
    file_name = "unknown_file"
    if msg.document:
        tg_file = await msg.document.get_file()
        file_name = msg.document.file_name or f"file_{len(context.user_data['files'])+1}"
    elif msg.photo:
        photo = msg.photo[-1]   # highest resolution
        tg_file = await photo.get_file()
        file_name = f"photo_{len(context.user_data['files'])+1}.jpg"
    elif msg.video:
        tg_file = await msg.video.get_file()
        file_name = msg.video.file_name or f"video_{len(context.user_data['files'])+1}"
    elif msg.audio:
        tg_file = await msg.audio.get_file()
        file_name = msg.audio.file_name or f"audio_{len(context.user_data['files'])+1}"

    if tg_file:
        dest = save_subdir / file_name
        await tg_file.download_to_drive(str(dest))
        context.user_data["files"].append(file_name)
        await msg.reply_text(
            f"📎 *{file_name}* хадгалагдлаа.\n\nӨөр файл илгээх, эсвэл дуусвал *Дууслаа* дарна уу.",
            parse_mode="Markdown",
        )
    else:
        await msg.reply_text("Файлыг уншиж чадсангүй. Баримт бичгийн хэлбэрээр илгээж үзнэ үү.")

    return AWAITING_FILES


async def finish_update(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """User typed 'done' or tapped the Done button — save and confirm."""
    save_update_to_excel(context.user_data)

    files = context.user_data.get("files", [])
    files_summary = f"{len(files)} файл: {', '.join(files)}" if files else "Байхгүй"

    await update.message.reply_text(
        "🎉 *Тайлан илгээгдлээ — баярлалаа!*\n\n"
        f"✅ Дуусгасан: {context.user_data.get('completed', '')[:80]}…\n"
        f"🔨 Ажиллаж байгаа: {context.user_data.get('working_on', '')[:80]}…\n"
        f"📎 Файлууд: {files_summary}\n\n"
        "_Таны тайлан долоо хоногийн хүснэгтэд хадгалагдлаа._",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove(),
    )
    context.user_data.clear()
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "Тайлан цуцлагдлаа. Бэлэн болсон үед /update ашиглана уу.",
        reply_markup=ReplyKeyboardRemove(),
    )
    context.user_data.clear()
    return ConversationHandler.END


# ── Basic commands ────────────────────────────────────────────────────────────

async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user = update.effective_user
    register_user(user.id, user.full_name, user.username)
    await update.message.reply_text(
        f"👋 Сайн байна уу, *{user.first_name}!*\n\n"
        "Та долоо хоногийн тайлангийн системд бүртгэгдлээ.\n\n"
        "Даваа гараг бүрийн өглөө би танаас гурван асуулт асуух болно:\n"
        "  1️⃣ Энэ долоо хоногт юу хийж дуусгасан бэ?\n"
        "  2️⃣ Одоо юу дээр ажиллаж байна вэ?\n"
        "  3️⃣ Илгээх файл байна уу?\n\n"
        "Та /update командаар хэдийд ч тайлан илгээж болно.\n\n"
        "Бүх командыг харахын тулд /help гэж бичнэ үү.",
        parse_mode="Markdown",
    )


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "*Боломжтой командууд:*\n\n"
        "/start — Бүртгүүлж эхлэх\n"
        "/update — Долоо хоногийн тайланг одоо илгээх\n"
        "/status — Хэдэн тайлан хадгалагдсаныг шалгах\n"
        "/sendreport — (Админ) Тайланг имэйлээр илгээх\n"
        "/promptall — (Админ) Бүх хэрэглэгчид сануулга илгээх\n"
        "/cancel — Тайлан илгээхийг цуцлах\n"
        "/help — Энэ мэдээллийг харуулах",
        parse_mode="Markdown",
    )


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    users = load_users()
    count = 0
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
        count = ws.max_row - 1   # subtract header row

    await update.message.reply_text(
        f"📊 *Байдал*\n\n"
        f"Бүртгэлтэй хэрэглэгчид: {len(users)}\n"
        f"Нийт хадгалсан тайлан: {count}",
        parse_mode="Markdown",
    )


async def cmd_sendreport(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Admin command — send the email report immediately."""
    await update.message.reply_text("Тайлангийн имэйл илгээж байна…")
    success, reason = send_weekly_email()
    if success:
        await update.message.reply_text(f"✅ Тайлан {MANAGER_EMAIL} руу илгээгдлээ")
    else:
        await update.message.reply_text(
            f"❌ Илгээхэд алдаа гарлаа. {EMAIL_ERROR_MESSAGES.get(reason, 'Тодорхойгүй алдаа.')}"
        )


async def cmd_promptall(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Admin command — send the weekly prompt to all users right now."""
    users = load_users()
    sent, failed = 0, 0
    for uid, info in users.items():
        if int(uid) == update.effective_user.id:
            continue   # don't message yourself
        try:
            await context.bot.send_message(
                chat_id=int(uid),
                text="👋 Таны менежер долоо хоногийн тайланг хүсч байна. /update командаар илгээнэ үү.",
            )
            sent += 1
        except Exception:
            failed += 1
    await update.message.reply_text(
        f"✅ {sent} хэрэглэгчид сануулга илгээгдлээ. Амжилтгүй: {failed}."
    )


# ── Entry point ───────────────────────────────────────────────────────────────

def main() -> None:
    if not BOT_TOKEN:
        raise SystemExit("BOT_TOKEN environment variable is not set.")

    _ensure_excel()

    app = Application.builder().token(BOT_TOKEN).build()

    # Weekly update conversation
    conv = ConversationHandler(
        entry_points=[CommandHandler("update", cmd_update)],
        states={
            COMPLETED: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_completed)
            ],
            WORKING_ON: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, receive_working_on)
            ],
            AWAITING_FILES: [
                MessageHandler(
                    filters.Document.ALL | filters.PHOTO | filters.VIDEO | filters.AUDIO,
                    receive_file,
                ),
                MessageHandler(
                    filters.Regex(r"(?i)^done") | filters.Regex(r"(?i)^no files") | filters.Regex(r"^Дууслаа"),
                    finish_update,
                ),
                MessageHandler(filters.TEXT & ~filters.COMMAND, finish_update),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("status", cmd_status))
    app.add_handler(CommandHandler("sendreport", cmd_sendreport))
    app.add_handler(CommandHandler("promptall", cmd_promptall))
    app.add_handler(conv)

    # Schedule jobs using PTB's built-in JobQueue
    import datetime as dt

    jq = app.job_queue
    # Monday 9:00 AM — send update prompts
    jq.run_daily(
        job_send_weekly_prompt,
        time=dt.time(PROMPT_HOUR, PROMPT_MINUTE, tzinfo=dt.timezone.utc),
        days=(PROMPT_DAY,),
        name="weekly_prompt",
    )
    # Friday 5:00 PM — email the Excel report
    jq.run_daily(
        job_send_weekly_report,
        time=dt.time(REPORT_HOUR, REPORT_MINUTE, tzinfo=dt.timezone.utc),
        days=(REPORT_DAY,),
        name="weekly_report",
    )

    logger.info("Bot started. Polling for messages…")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
